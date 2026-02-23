import os
from io import StringIO, BytesIO
import csv
from flask import Flask, render_template, request, redirect, session, url_for, flash, send_file
from werkzeug.security import check_password_hash
from auth import login_required, require_role
from psycopg import errors
from werkzeug.utils import secure_filename
import tempfile
from openpyxl import load_workbook, Workbook
from services import (
    get_employees, get_projects, get_project_assignments,
    upsert_hours, add_employee, update_employee, delete_employee,
    get_managers_summary
)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev")

# authentication
from db import get_conn  # only used for login lookup

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        u = request.form["username"]; p = request.form["password"]
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("SELECT id, password_hash, role FROM app_user WHERE username=%s", (u,))
            row = cur.fetchone()
        if row and check_password_hash(row["password_hash"], p):
            session["user_id"] = row["id"]; session["role"] = row["role"]
            return redirect(url_for("employees"))
        flash("Invalid username or password")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("login"))

@app.route("/forbidden")
def forbidden():
    return "Forbidden (viewer cannot perform this action)", 403

# ORDER BY allow-lists (safe sorting)
EMPLOYEE_ORDER = {
    "name_asc":  "full_name ASC",
    "name_desc": "full_name DESC",
    "hours_asc": "total_hours ASC",
    "hours_desc":"total_hours DESC",
}
PROJECT_ORDER = {
    "headcount_asc":  "headcount ASC",
    "headcount_desc": "headcount DESC",
    "hours_asc":      "total_hours ASC",
    "hours_desc":     "total_hours DESC",
}

# Employees (A2)
@app.route("/", methods=["GET"])
@login_required
def employees():
    dept = request.args.get("dept")
    q    = request.args.get("q")
    sort = request.args.get("sort","name_asc")

    # sanitize dept: digits only, else treat as empty
    if dept is not None:
        dept = dept.strip()
        if dept == "" or not dept.isdigit():
            if dept:  # user typed something non-numeric
                flash("Dept # must be a number.")
            dept = None

    q = (q or "").strip() or None

    order_sql = EMPLOYEE_ORDER.get(sort, "full_name ASC")
    rows = get_employees(dept, q, order_sql)
    return render_template("employees.html",
                           rows=rows, dept=dept, q=q, sort=sort,
                           order_keys=list(EMPLOYEE_ORDER.keys()))



# Projects (A3) 
@app.route("/projects", methods=["GET"])
@login_required
def projects():
    sort = request.args.get("sort","headcount_desc")
    order_sql = PROJECT_ORDER.get(sort, "headcount DESC")
    rows = get_projects(order_sql)
    return render_template("projects.html", rows=rows, sort=sort, order_keys=list(PROJECT_ORDER.keys()))

# Project Details + Add Hours (A4)
@app.route("/projects/<int:pno>", methods=["GET","POST"])
@login_required
def project_detail(pno:int):
    msg = None

    if request.method == "POST":
        if session.get("role") != "admin":
            return redirect(url_for("forbidden"))  # 403 page

        essn  = request.form.get("essn","").strip()
        hours_raw = request.form.get("hours","").strip()

        # validate hours number > 0
        try:
            hours = float(hours_raw)
            if hours <= 0:
                raise ValueError
        except Exception:
            flash("Hours must be a positive number.")
            return redirect(url_for("project_detail", pno=pno))

        # validate project + employee exist (avoid FK crash)
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("SELECT 1 FROM Project WHERE Pnumber=%s", (pno,))
            if not cur.fetchone():
                flash("Project not found.")
                return redirect(url_for("projects"))

            cur.execute("SELECT 1 FROM Employee WHERE Ssn=%s", (essn,))
            if not cur.fetchone():
                flash("No such employee SSN.")
                return redirect(url_for("project_detail", pno=pno))

        # do the upsert
        try:
            upsert_hours(essn, pno, hours)
            flash("Hours recorded.")
        except Exception as e:
            # last-ditch safety if DB rejects insert
            flash(f"Could not record hours: {e}")

    rows = get_project_assignments(pno)
    # also fetch employees for the dropdown (ssn + full_name)
    employees_list = get_employees(None, None, "full_name ASC")
    return render_template("project_detail.html", pno=pno, rows=rows, msg=msg, employees=employees_list)


# Employees CRUD (A5) 
@app.route("/employees/new", methods=["GET","POST"])
@login_required
@require_role('admin')
def employee_new():
    if request.method == "POST":
        ssn   = request.form["ssn"].strip()
        fname = request.form["fname"].strip()
        minit = (request.form.get("minit","-") or "-")[:1].upper()
        lname = request.form["lname"].strip()
        addr  = request.form["address"].strip()
        sex   = request.form["sex"]
        sal_raw = request.form["salary"].strip()
        dno_raw = request.form["dno"].strip()
        super_ssn = request.form.get("super_ssn") or None

        # basic server-side validation
        if not ssn.isdigit():
            flash("SSN must be digits only."); return render_template("employee_form.html")
        try:
            sal = float(sal_raw);  dno = int(dno_raw)
            if sal < 0 or dno < 1: raise ValueError
        except ValueError:
            flash("Salary must be ≥ 0 and Dept # must be a positive integer.")
            return render_template("employee_form.html")

        try:
            add_employee(ssn, fname, minit, lname, addr, sex, sal, dno, super_ssn)
            flash("Employee added.")
            return redirect(url_for("employees"))
        except errors.UniqueViolation:
            flash("That SSN already exists. Use a different SSN.")
            return render_template("employee_form.html")
        except Exception as e:
            flash(f"Could not add employee: {e}")
            return render_template("employee_form.html")

    return render_template("employee_form.html")




@app.route("/employees/<ssn>/edit", methods=["GET","POST"])
@login_required
@require_role('admin')
def employee_edit(ssn):
    if request.method == "POST":
        addr  = request.form["address"].strip()
        sal   = float(request.form["salary"])
        dno   = int(request.form["dno"])
        super_ssn = request.form.get("super_ssn") or None
        sex   = request.form.get("sex")                 
        minit = (request.form.get("minit","-") or "-")[:1].upper()

        update_employee(ssn, addr, sal, dno, super_ssn, sex, minit)
        flash("Employee updated.")
        return redirect(url_for("employees"))

    # GET pull current values to prefill
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT Ssn AS ssn, Address AS address, Salary AS salary, Dno AS dno,
                   Sex AS sex, Minit AS minit, Super_ssn AS super_ssn
            FROM Employee WHERE Ssn=%s
        """, (ssn,))
        row = cur.fetchone()
    return render_template("employee_form.html", ssn=ssn, row=row)

@app.route("/employees/<ssn>/delete", methods=["POST"])
@login_required
@require_role('admin')
def employee_delete_route(ssn):
    try:
        delete_employee(ssn)
        flash("Employee deleted.")
    except ValueError as e:
        flash(str(e))
    return redirect(url_for('employees'))

#Managers overview (A6) 
@app.route("/managers")
@login_required
def managers():
    rows = get_managers_summary()
    return render_template("managers.html", rows=rows)

#CSV export
@app.route("/employees.csv")
@login_required
def employees_csv():
    dept = request.args.get("dept")
    q    = request.args.get("q")
    sort = request.args.get("sort","name_asc")

    dept = None if (dept is None or dept.strip() == "") else dept
    q    = None if (q is None or q.strip() == "") else q

    order_sql = EMPLOYEE_ORDER.get(sort, "full_name ASC")
    rows = get_employees(dept, q, order_sql)

    # build CSV in text
    sio = StringIO()
    w = csv.writer(sio)
    w.writerow(["Full Name","Department","#Dependents","#Projects","Total Hours"])
    for r in rows:
        w.writerow([
            r["full_name"],
            r["dept_name"] or "N/A",
            r["num_dependents"],
            r["num_projects"],
            f'{(r["total_hours"] or 0):.2f}',
        ])

    # convert to bytes for send_file
    data = sio.getvalue().encode("utf-8")
    bio = BytesIO(data)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name="employees.csv",
    )

@app.route("/projects/<int:pno>/import-template.xlsx")
@login_required
@require_role('admin')
def project_import_template(pno:int):
    # Create a tiny Excel with header: SSN, Hours
    wb = Workbook()
    ws = wb.active
    ws.title = "Hours"
    ws.append(["SSN", "Hours"])   # header
    # (optional example rows)
    ws.append(["200243095", 8.0])
    ws.append(["359624751", 3.5])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name); tmp.seek(0)

    return send_file(tmp.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"project_{pno}_hours_template.xlsx")


@app.route("/projects/<int:pno>/import", methods=["POST"])
@login_required
@require_role('admin')
def project_import_hours(pno:int):
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Please choose an .xlsx file.")
        return redirect(url_for("project_detail", pno=pno))

    # Save to a temp file so openpyxl can read it
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    file.save(tmp.name)

    try:
        wb = load_workbook(tmp.name, data_only=True)
        ws = wb.active
    except Exception:
        flash("Could not read Excel file (expect .xlsx).")
        return redirect(url_for("project_detail", pno=pno))

    rows_ok, errors_list = 0, []
    # Expect header in row 1: SSN, Hours
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        essn, hours = (row + (None, None))[:2]
        # Validate
        if essn is None or hours is None:
            errors_list.append(f"Row {i}: missing SSN or Hours"); continue
        essn = str(essn).strip()
        try:
            hours = float(hours)
            if hours <= 0: raise ValueError
        except Exception:
            errors_list.append(f"Row {i}: Hours must be a positive number"); continue
        if not essn.isdigit():
            errors_list.append(f"Row {i}: SSN must be digits only"); continue

        # Check employee exists (avoid FK failure)
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("SELECT 1 FROM Employee WHERE Ssn=%s", (essn,))
            if not cur.fetchone():
                errors_list.append(f"Row {i}: No employee with SSN {essn}")
                continue

        # Upsert
        try:
            upsert_hours(essn, pno, hours)
            rows_ok += 1
        except Exception as e:
            errors_list.append(f"Row {i}: DB error: {e}")

    if rows_ok:
        flash(f"Imported {rows_ok} row(s) into project {pno}.")
    if errors_list:
        # show only the first few to keep UI tidy
        preview = errors_list[:5]
        more = "" if len(errors_list) <= 5 else f" (+{len(errors_list)-5} more)"
        flash("Some rows were skipped: " + "; ".join(preview) + more)

    return redirect(url_for("project_detail", pno=pno))
